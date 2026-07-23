import random
import os
import sys
import csv
import io
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import uuid
import threading
import queue
import sqlite3 as sqlite3_module
from datetime import datetime, timezone, timedelta

# 北京时间
TZ_CN = timezone(timedelta(hours=8))

def now_cn():
    """返回北京时间 datetime"""
    return datetime.now(TZ_CN).replace(tzinfo=None)

def now_cn_str(fmt="%Y-%m-%d %H:%M:%S"):
    """返回北京时间字符串"""
    return datetime.now(TZ_CN).strftime(fmt)

try:
    import pymysql
    import pymysql.cursors
    HAS_PYMYSQL = True
except ImportError:
    HAS_PYMYSQL = False
from flask import Flask, g, render_template, request, jsonify, redirect, url_for, session, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

from config import Config


app = Flask(__name__)
app.config.from_object(Config)

# Session 过期配置
app.config["PERMANENT_SESSION_LIFETIME"] = Config.PERMANENT_SESSION_LIFETIME

# 静态文件版本号
APP_VERSION = str(int(os.path.getmtime(__file__)))


@app.context_processor
def inject_version():
    return dict(app_version=APP_VERSION)


# ═══════════════════════════════════════════
# 数据库（MySQL / SQLite 自适应）
# ═══════════════════════════════════════════

DB_IS_MYSQL = (Config.DB_TYPE == "mysql")


class DB:
    """统一数据库封装，兼容 MySQL pymysql cursor 和 SQLite cursor"""
    def __init__(self, conn, cursor):
        self._conn = conn
        self._cursor = cursor

    def execute(self, sql, params=None):
        # MySQL 用 %s 占位符，SQLite 用 ?，这里统一转换
        if not DB_IS_MYSQL:
            sql = sql.replace("%s", "?")
        self._cursor.execute(sql, params or ())
        return self._cursor

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    @property
    def rowcount(self):
        return self._cursor.rowcount

    @property
    def lastrowid(self):
        return self._cursor.lastrowid

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._cursor.close()
        self._conn.close()


def get_db():
    """获取数据库连接（请求级别复用）"""
    if "db" not in g:
        if DB_IS_MYSQL:
            conn = pymysql.connect(
                host=app.config["MYSQL_HOST"],
                port=app.config["MYSQL_PORT"],
                user=app.config["MYSQL_USER"],
                password=app.config["MYSQL_PASSWORD"],
                database=app.config["MYSQL_DATABASE"],
                charset="utf8mb4",
                cursorclass=pymysql.cursors.DictCursor,
                autocommit=False,
            )
            g.db = DB(conn, conn.cursor())
        else:
            conn = sqlite3_module.connect(Config.DATABASE)
            conn.row_factory = sqlite3_module.Row
            conn.execute("PRAGMA journal_mode=WAL")
            g.db = DB(conn, conn.cursor())
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        try:
            if exception:
                db.rollback()
            db.close()
        except Exception:
            pass


def init_db():
    """初始化数据库表结构并执行自动迁移（MySQL / SQLite 自适应）"""
    if DB_IS_MYSQL:
        db = pymysql.connect(
            host=app.config["MYSQL_HOST"],
            port=app.config["MYSQL_PORT"],
            user=app.config["MYSQL_USER"],
            password=app.config["MYSQL_PASSWORD"],
            database=app.config["MYSQL_DATABASE"],
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=True,
        )
        cur = db.cursor()
        engine_sql = "ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
        int_pk = "INT AUTO_INCREMENT PRIMARY KEY"
        varchar = "VARCHAR"
        timestamp_default = "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        tinyint = "TINYINT"
        text_type = "TEXT"
    else:
        db = sqlite3_module.connect(Config.DATABASE)
        db.row_factory = sqlite3_module.Row
        db.execute("PRAGMA journal_mode=WAL")
        cur = db.cursor()
        engine_sql = ""
        int_pk = "INTEGER PRIMARY KEY AUTOINCREMENT"
        varchar = "TEXT"
        timestamp_default = "TIMESTAMP DEFAULT (datetime('now', '+8 hours'))"
        tinyint = "INTEGER"
        text_type = "TEXT"

    def _exec(sql, params=None):
        # 替换 MySQL 专有语法
        sql = sql.replace("ENGINE=InnoDB DEFAULT CHARSET=utf8mb4", engine_sql)
        if not DB_IS_MYSQL:
            sql = sql.replace("INT AUTO_INCREMENT PRIMARY KEY", int_pk)
            sql = sql.replace("VARCHAR", varchar)
            sql = sql.replace("TINYINT", tinyint)
            sql = sql.replace("TEXT", text_type)
            sql = sql.replace("%s", "?")
            # SQLite 不支持 ALTER TABLE ADD COLUMN 后的 AFTER
            import re
            sql = re.sub(r'\s+AFTER\s+\w+', '', sql, flags=re.IGNORECASE)
        cur.execute(sql, params or ())
        if not DB_IS_MYSQL:
            db.commit()

    _exec(f"""
        CREATE TABLE IF NOT EXISTS admissions (
            id {int_pk},
            name {varchar}(64) NOT NULL,
            category {varchar}(128) DEFAULT '',
            class_type {varchar}(16) DEFAULT 'kete',
            grade {varchar}(16) DEFAULT '',
            score {varchar}(16) DEFAULT '',
            created_at {timestamp_default}
        ) {engine_sql}
    """)

    _exec(f"""
        CREATE TABLE IF NOT EXISTS query_logs (
            id {int_pk},
            name {varchar}(64) NOT NULL,
            admitted {tinyint} DEFAULT 0,
            class_type {varchar}(16) DEFAULT '',
            needs {text_type},
            created_at {timestamp_default}
        ) {engine_sql}
    """)

    _exec(f"""
        CREATE TABLE IF NOT EXISTS admin_audit_log (
            id {int_pk},
            action {varchar}(64) NOT NULL,
            target {varchar}(128) DEFAULT '',
            detail {varchar}(512) DEFAULT '',
            admin_user {varchar}(32) DEFAULT 'admin',
            created_at {timestamp_default}
        ) {engine_sql}
    """)

    _exec(f"""
        CREATE TABLE IF NOT EXISTS captcha_store (
            id {int_pk},
            token {varchar}(32) NOT NULL UNIQUE,
            answer INT NOT NULL,
            used {tinyint} DEFAULT 0,
            created_at {timestamp_default}
        ) {engine_sql}
    """)

    _exec(f"""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id {int_pk},
            ip_address {varchar}(64) NOT NULL,
            success {tinyint} DEFAULT 0,
            created_at {timestamp_default}
        ) {engine_sql}
    """)

    _exec(f"""
        CREATE TABLE IF NOT EXISTS schema_version (
            version INT PRIMARY KEY,
            applied_at {timestamp_default}
        ) {engine_sql}
    """)

    # 获取当前 schema 版本
    db_wrapper = DB(db, cur)
    db_wrapper.execute("SELECT MAX(version) as v FROM schema_version")
    row = db_wrapper.fetchone()
    current_version = row["v"] if row and row["v"] is not None else 0

    # 创建索引（SQLite 版本）
    if not DB_IS_MYSQL:
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_admissions_name ON admissions(name)",
            "CREATE INDEX IF NOT EXISTS idx_admissions_class ON admissions(class_type)",
            "CREATE INDEX IF NOT EXISTS idx_query_logs_name ON query_logs(name)",
            "CREATE INDEX IF NOT EXISTS idx_query_logs_class ON query_logs(class_type)",
            "CREATE INDEX IF NOT EXISTS idx_query_logs_created ON query_logs(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_captcha_token ON captcha_store(token)",
            "CREATE INDEX IF NOT EXISTS idx_login_ip ON login_attempts(ip_address, created_at)",
        ]:
            try:
                cur.execute(idx_sql)
                db.commit()
            except Exception:
                pass

    # 迁移脚本
    migrations = {
        1: [
            "ALTER TABLE admissions ADD COLUMN class_type VARCHAR(16) DEFAULT 'kete'",
            "ALTER TABLE query_logs ADD COLUMN class_type VARCHAR(16) DEFAULT ''",
        ],
        2: [
            f"""CREATE TABLE IF NOT EXISTS admin_audit_log (
                id {int_pk},
                action {varchar}(64) NOT NULL,
                target {varchar}(128) DEFAULT '',
                detail {varchar}(512) DEFAULT '',
                admin_user {varchar}(32) DEFAULT 'admin',
                created_at {timestamp_default}
            ) {engine_sql}""",
            f"""CREATE TABLE IF NOT EXISTS captcha_store (
                id {int_pk},
                token {varchar}(32) NOT NULL UNIQUE,
                answer INT NOT NULL,
                used {tinyint} DEFAULT 0,
                created_at {timestamp_default}
            ) {engine_sql}""",
        ],
        3: [],
        4: [
            "ALTER TABLE query_logs ADD COLUMN needs TEXT",
        ],
        5: [
            f"""CREATE TABLE IF NOT EXISTS login_attempts (
                id {int_pk},
                ip_address {varchar}(64) NOT NULL,
                success {tinyint} DEFAULT 0,
                created_at {timestamp_default}
            ) {engine_sql}""",
        ],
        6: [
            # 将旧 UTC 时间转换为北京时间（+8小时）
            "UPDATE admissions SET created_at = datetime(created_at, '+8 hours') WHERE created_at IS NOT NULL",
            "UPDATE query_logs SET created_at = datetime(created_at, '+8 hours') WHERE created_at IS NOT NULL",
            "UPDATE admin_audit_log SET created_at = datetime(created_at, '+8 hours') WHERE created_at IS NOT NULL",
            "UPDATE captcha_store SET created_at = datetime(created_at, '+8 hours') WHERE created_at IS NOT NULL",
            "UPDATE login_attempts SET created_at = datetime(created_at, '+8 hours') WHERE created_at IS NOT NULL",
        ],
        7: [
            # 添加 IP 地址字段，用于查询频率限制
            "ALTER TABLE query_logs ADD COLUMN ip_address VARCHAR(64) DEFAULT ''",
        ],
    }

    for ver in sorted(migrations.keys()):
        if ver <= current_version:
            continue
        for sql in migrations[ver]:
            if not sql:
                continue
            try:
                _exec(sql)
            except Exception as e:
                err_msg = str(e)
                if "Duplicate column" in err_msg or "already exists" in err_msg or "duplicate column name" in err_msg.lower():
                    pass
                else:
                    print(f"[DB Migration] v{ver} error: {e}")
        _exec("INSERT INTO schema_version (version) VALUES (%s)", (ver,))
        print(f"[DB Migration] v{ver} applied")

    cur.close()
    if not DB_IS_MYSQL:
        db.close()
    else:
        db.close()
    print(f"[DB] Initialized successfully (type={Config.DB_TYPE})")


# 应用启动时初始化数据库
init_db()


# ═══════════════════════════════════════════
# 异步日志写入队列（高并发��化）
# ═══════════════════════════════════════════

_log_queue = queue.Queue()


def _log_writer_thread():
    """后台线程：从队列取出日志写入数据库，批量提交"""
    if DB_IS_MYSQL:
        conn = pymysql.connect(
            host=app.config["MYSQL_HOST"],
            port=app.config["MYSQL_PORT"],
            user=app.config["MYSQL_USER"],
            password=app.config["MYSQL_PASSWORD"],
            database=app.config["MYSQL_DATABASE"],
            charset="utf8mb4",
            autocommit=False,
        )
    else:
        conn = sqlite3_module.connect(Config.DATABASE)
        conn.row_factory = sqlite3_module.Row
        conn.execute("PRAGMA journal_mode=WAL")
    cur = conn.cursor()
    batch = []
    last_flush = time.time()

    while True:
        try:
            item = _log_queue.get(timeout=2)
            if item is None:
                break
            batch.append(item)

            if len(batch) >= 50 or (time.time() - last_flush > 1 and batch):
                _flush_batch(cur, conn, batch)
                batch = []
                last_flush = time.time()
        except queue.Empty:
            if batch:
                _flush_batch(cur, conn, batch)
                batch = []
                last_flush = time.time()

    if batch:
        _flush_batch(cur, conn, batch)
    cur.close()
    conn.close()


def _flush_batch(cur, conn, batch):
    """批量写入日志"""
    try:
        if DB_IS_MYSQL:
            cur.executemany(
                "INSERT INTO query_logs (name, admitted, class_type) VALUES (%s, %s, %s)",
                batch,
            )
        else:
            cur.executemany(
                "INSERT INTO query_logs (name, admitted, class_type) VALUES (?, ?, ?)",
                batch,
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


# 启动后台日志写入线程
_log_thread = threading.Thread(target=_log_writer_thread, daemon=True)
_log_thread.start()


# ═══════════════════════════════════════════
# 班型配置
# ═══════════════════════════════════════════

CLASS_TYPES = {
    "kete": {"name": "科特班", "title": "科特班·英才计划录取结果查询", "category": "科特班·英才计划"},
    "yucai": {"name": "育才班", "title": "育才班·英才计划录取结果查询", "category": "育才班·英才计划"},
}


# ═══════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════

def generate_grade():
    pct = random.randint(1, 9)
    return f"前{pct}%"


def generate_captcha():
    ops = [
        ("+", lambda a, b: a + b),
        ("-", lambda a, b: a - b),
        ("×", lambda a, b: a * b),
    ]
    op_symbol, op_func = random.choice(ops)
    if op_symbol == "×":
        a, b = random.randint(1, 9), random.randint(1, 9)
    elif op_symbol == "-":
        a, b = random.randint(5, 20), random.randint(1, 10)
        if a < b:
            a, b = b, a
    else:
        a, b = random.randint(1, 20), random.randint(1, 10)
    answer = op_func(a, b)
    expression = f"{a} {op_symbol} {b} = ?"
    return expression, answer


def log_audit(action, target="", detail=""):
    try:
        if DB_IS_MYSQL:
            conn = pymysql.connect(
                host=app.config["MYSQL_HOST"],
                port=app.config["MYSQL_PORT"],
                user=app.config["MYSQL_USER"],
                password=app.config["MYSQL_PASSWORD"],
                database=app.config["MYSQL_DATABASE"],
                charset="utf8mb4",
                autocommit=True,
            )
        else:
            conn = sqlite3_module.connect(Config.DATABASE)
            conn.row_factory = sqlite3_module.Row
            conn.execute("PRAGMA journal_mode=WAL")
        cur = conn.cursor()
        if not DB_IS_MYSQL:
            cur.execute(
                "INSERT INTO admin_audit_log (action, target, detail, created_at) VALUES (?, ?, ?, ?)",
                (action, target, detail, now_cn_str()),
            )
        else:
            cur.execute(
                "INSERT INTO admin_audit_log (action, target, detail, created_at) VALUES (%s, %s, %s, %s)",
                (action, target, detail, now_cn_str()),
            )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[Audit] log_audit error: {e}", file=sys.stderr)
        pass


# ═══════════════════════════════════════════
# 认证
# ═══════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login_page"))
        return f(*args, **kwargs)
    return decorated


def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return jsonify({"success": False, "message": "未登录，请先登录", "code": 401}), 401
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════
# 前端路由
# ═══════════════════════════════════════════

@app.route("/")
def root():
    return redirect("/kete")


@app.route("/admin")
def admin_login_page():
    return render_template("admin_login.html")


@app.route("/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json()
    username = data.get("username", "")
    password = data.get("password", "")
    client_ip = request.remote_addr or "unknown"

    db = get_db()

    failed_count = 0
    if DB_IS_MYSQL:
        row = db.execute(
            """SELECT COUNT(*) as cnt FROM login_attempts
               WHERE ip_address = %s AND success = 0
               AND created_at > DATE_SUB(NOW(), INTERVAL %s MINUTE)""",
            (client_ip, app.config["LOGIN_LOCKOUT_MINUTES"]),
        ).fetchone()
    else:
        row = db.execute(
            """SELECT COUNT(*) as cnt FROM login_attempts
               WHERE ip_address = %s AND success = 0
               AND created_at > datetime('now', '-' || %s || ' minutes')""",
            (client_ip, str(app.config["LOGIN_LOCKOUT_MINUTES"])),
        ).fetchone()
    if row:
        failed_count = row["cnt"]

    if failed_count >= app.config["MAX_LOGIN_ATTEMPTS"]:
        return jsonify({
            "success": False,
            "message": f"登录尝试次数过多，请 {app.config['LOGIN_LOCKOUT_MINUTES']} 分钟后再试"
        })

    if username == app.config["ADMIN_USERNAME"] and check_password_hash(app.config["ADMIN_PASSWORD_HASH"], password):
        session["admin_logged_in"] = True
        session.permanent = True
        db.execute(
            "INSERT INTO login_attempts (ip_address, success, created_at) VALUES (%s, 1, %s)",
            (client_ip, now_cn_str()),
        )
        db.commit()
        return jsonify({"success": True})

    db.execute(
        "INSERT INTO login_attempts (ip_address, success, created_at) VALUES (%s, 0, %s)",
        (client_ip, now_cn_str()),
    )
    db.commit()
    time.sleep(1)
    return jsonify({"success": False, "message": "账号或密码错误"})


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login_page"))


@app.route("/admin/dashboard")
@login_required
def admin_dashboard():
    return render_template("admin_dashboard.html")


@app.route("/<class_type>")
def index(class_type):
    if class_type not in CLASS_TYPES:
        return render_template("error.html", code=404, message="页面不存在"), 404
    ct = CLASS_TYPES[class_type]
    return render_template("index.html", class_type=class_type, class_name=ct["name"], page_title=ct["title"])


@app.route("/<class_type>/result")
def result_page(class_type):
    if class_type not in CLASS_TYPES:
        return render_template("error.html", code=404, message="页面不存在"), 404
    ct = CLASS_TYPES[class_type]
    name = request.args.get("name", "")
    category = request.args.get("category", ct["category"])
    grade = request.args.get("grade", "")
    score = request.args.get("score", "")
    return render_template("result.html",
        class_type=class_type,
        class_name=ct["name"],
        category=category,
        name=name,
        grade=grade,
        score=score)


@app.route("/<class_type>/invite")
def invite_page(class_type):
    if class_type not in CLASS_TYPES:
        return render_template("error.html", code=404, message="页面不存在"), 404
    ct = CLASS_TYPES[class_type]
    student_name = request.args.get("name", "")
    badge_text = "英才计划录取资格"
    today = now_cn_str("%Y年%m月%d日")
    return render_template("invite.html",
        class_type=class_type,
        class_name=ct["name"],
        student_name=student_name,
        badge_text=badge_text,
        today_date=today)


# ═══════════════════════════════════════════
# 验证码 API
# ═══════════════════════════════════════════

@app.route("/api/captcha")
def get_captcha():
    db = get_db()
    # 自动清理超过10分钟的未使用验证码
    if DB_IS_MYSQL:
        db.execute("DELETE FROM captcha_store WHERE used = 0 AND created_at < DATE_SUB(NOW(), INTERVAL 10 MINUTE)")
    else:
        db.execute("DELETE FROM captcha_store WHERE used = 0 AND created_at < datetime('now', '-10 minutes')")
    db.commit()

    expression, answer = generate_captcha()
    token = uuid.uuid4().hex[:16]

    db.execute(
        "INSERT INTO captcha_store (token, answer, created_at) VALUES (%s, %s, %s)",
        (token, answer, now_cn_str()),
    )
    db.commit()
    return jsonify({"token": token, "expression": expression})


# ═══════════════════════════════════════════
# 查询 API（异步日志写入，高并发优化）
# ═══════════════════════════════════════════

@app.route("/api/query")
def query():
    name = request.args.get("name", "").strip()
    class_type = request.args.get("class_type", "kete").strip()
    captcha_token = request.args.get("captcha_token", "").strip()
    captcha_answer = request.args.get("captcha_answer", "").strip()

    if not name:
        return jsonify({"success": False, "message": "请输入姓名"})

    db = get_db()

    # 验证码验证
    if captcha_token and captcha_answer:
        captcha_row = db.execute(
            "SELECT answer, used FROM captcha_store WHERE token = %s",
            (captcha_token,),
        ).fetchone()
        if not captcha_row:
            return jsonify({"success": False, "message": "验证码已过期，请刷新后重试"})
        if captcha_row["used"]:
            return jsonify({"success": False, "message": "验证码已使用，请刷新后重试"})
        try:
            if int(captcha_answer) != captcha_row["answer"]:
                return jsonify({"success": False, "message": "验证码错误，请重新输入"})
        except (ValueError, TypeError):
            return jsonify({"success": False, "message": "验证码格式错误"})
        # 原子标记已使用（高并发安全）
        affected = db.execute(
            "UPDATE captcha_store SET used = 1 WHERE token = %s AND used = 0",
            (captcha_token,),
        )
        if affected == 0:
            return jsonify({"success": False, "message": "验证码已使用，请刷新后重试"})
        db.commit()
    else:
        return jsonify({"success": False, "message": "请输入验证码"})

    # 查询频率限制：同一 IP + 同一姓名 30 分钟内最多 3 次
    client_ip = request.remote_addr or "unknown"
    rate_limit_row = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE name = %s AND ip_address = %s AND created_at > datetime('now', '-30 minutes', '+8 hours')",
        (name, client_ip),
    ).fetchone()
    if rate_limit_row and rate_limit_row["cnt"] >= 3:
        return jsonify({"success": False, "message": "查询次数过多，请30分钟后再试"})

    # 查询录取结果
    row = db.execute(
        "SELECT name, category, class_type, grade, score FROM admissions WHERE name = %s AND class_type = %s",
        (name, class_type),
    ).fetchone()

    if row:
        grade = row["grade"] or generate_grade()
        if not row["grade"]:
            db.execute(
                "UPDATE admissions SET grade = %s WHERE name = %s AND class_type = %s",
                (grade, name, class_type),
            )
            db.commit()

        ct = CLASS_TYPES.get(class_type, CLASS_TYPES["kete"])
        category = row["category"] or ct["category"]

        # 同步写入查询日志（确保管理后台立即可见）
        db.execute(
            "INSERT INTO query_logs (name, admitted, class_type, ip_address, created_at) VALUES (%s, 1, %s, %s, %s)",
            (name, class_type, client_ip, now_cn_str()),
        )
        db.commit()

        return jsonify({
            "success": True,
            "admitted": True,
            "name": row["name"],
            "category": category,
            "class_type": class_type,
            "grade": grade,
            "score": ""
        })
    else:
        # 同步写入未录取日志
        db.execute(
            "INSERT INTO query_logs (name, admitted, class_type, ip_address, created_at) VALUES (%s, 0, %s, %s, %s)",
            (name, class_type, client_ip, now_cn_str()),
        )
        db.commit()

        return jsonify({"success": True, "admitted": False, "message": "未查询到录取信息"})


# ═══════════════════════════════════════════
# 管理后台 API
# ═══════════════════════════════════════════

@app.route("/api/admin/stats")
@api_login_required
def admin_stats():
    db = get_db()
    total = db.execute("SELECT COUNT(*) as cnt FROM admissions").fetchone()["cnt"]
    kete = db.execute("SELECT COUNT(*) as cnt FROM admissions WHERE class_type = 'kete'").fetchone()["cnt"]
    yucai = db.execute("SELECT COUNT(*) as cnt FROM admissions WHERE class_type = 'yucai'").fetchone()["cnt"]

    today = now_cn_str("%Y-%m-%d")
    today_queries = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE DATE(created_at) = %s",
        (today,),
    ).fetchone()["cnt"]

    total_queries = db.execute("SELECT COUNT(*) as cnt FROM query_logs").fetchone()["cnt"]
    confirmed = db.execute("SELECT COUNT(*) as cnt FROM query_logs WHERE needs != ''").fetchone()["cnt"]

    admitted_count = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE admitted = 1"
    ).fetchone()["cnt"]

    admission_rate = round(admitted_count / total_queries * 100, 1) if total_queries > 0 else 0

    today_new = db.execute(
        "SELECT COUNT(*) as cnt FROM admissions WHERE DATE(created_at) = %s",
        (today,),
    ).fetchone()["cnt"]

    kete_queries = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE class_type = 'kete'"
    ).fetchone()["cnt"]
    yucai_queries = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE class_type = 'yucai'"
    ).fetchone()["cnt"]
    kete_confirmed = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE class_type = 'kete' AND needs != ''"
    ).fetchone()["cnt"]
    yucai_confirmed = db.execute(
        "SELECT COUNT(*) as cnt FROM query_logs WHERE class_type = 'yucai' AND needs != ''"
    ).fetchone()["cnt"]

    return jsonify({
        "total": total,
        "kete": kete,
        "yucai": yucai,
        "today_queries": today_queries,
        "total_queries": total_queries,
        "confirmed": confirmed,
        "admission_rate": admission_rate,
        "today_new": today_new,
        "kete_queries": kete_queries,
        "yucai_queries": yucai_queries,
        "kete_confirmed": kete_confirmed,
        "yucai_confirmed": yucai_confirmed
    })


@app.route("/api/admin/upload", methods=["POST"])
@api_login_required
def admin_upload():
    data = request.get_json()
    if not data or "names" not in data:
        return jsonify({"success": False, "message": "未收到数据"})

    names = data["names"]
    class_type = data.get("class_type", "kete")

    if not names:
        return jsonify({"success": False, "message": "名单为空"})

    db = get_db()
    inserted = 0
    skipped = 0

    for item in names:
        name = item.get("name", "").strip() if isinstance(item, dict) else str(item).strip()
        category = item.get("category", "").strip() if isinstance(item, dict) else ""

        if not name:
            continue

        existing = db.execute(
            "SELECT id FROM admissions WHERE name = %s AND class_type = %s",
            (name, class_type),
        ).fetchone()

        if existing:
            skipped += 1
            continue

        try:
            db.execute(
                "INSERT INTO admissions (name, category, class_type, created_at) VALUES (%s, %s, %s, %s)",
                (name, category, class_type, now_cn_str()),
            )
            inserted += 1
        except Exception:
            pass

    db.commit()

    log_audit("上传录取名单", f"{CLASS_TYPES.get(class_type, {}).get('name', class_type)}", f"成功 {inserted} 条，跳过 {skipped} 条")

    return jsonify({
        "success": True,
        "inserted": inserted,
        "skipped": skipped,
    })


@app.route("/api/admin/list")
@api_login_required
def admin_list():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "").strip()
    filter_class = request.args.get("class_type", "").strip()

    db = get_db()

    conditions = []
    params = []
    if search:
        conditions.append("name LIKE %s")
        params.append(f"%{search}%")
    if filter_class:
        conditions.append("class_type = %s")
        params.append(filter_class)

    if conditions:
        where = " AND ".join(conditions)
        count = db.execute(
            f"SELECT COUNT(*) as cnt FROM admissions WHERE {where}", params
        ).fetchone()["cnt"]
        rows = db.execute(
            f"SELECT id, name, class_type, created_at FROM admissions WHERE {where} ORDER BY id DESC LIMIT %s OFFSET %s",
            params + [per_page, (page - 1) * per_page],
        ).fetchall()
    else:
        count = db.execute("SELECT COUNT(*) as cnt FROM admissions").fetchone()["cnt"]
        rows = db.execute(
            "SELECT id, name, class_type, created_at FROM admissions ORDER BY id DESC LIMIT %s OFFSET %s",
            (per_page, (page - 1) * per_page),
        ).fetchall()

    return jsonify({
        "total": count,
        "page": page,
        "per_page": per_page,
        "items": [dict(r) for r in rows] if rows else []
    })


@app.route("/api/admin/update/<int:record_id>", methods=["PUT"])
@api_login_required
def admin_update(record_id):
    data = request.get_json()
    new_name = data.get("name", "").strip()
    new_class_type = data.get("class_type", "").strip()

    if not new_name:
        return jsonify({"success": False, "message": "姓名不能为空"})
    if new_class_type not in CLASS_TYPES:
        return jsonify({"success": False, "message": "班型无效"})

    db = get_db()

    existing = db.execute(
        "SELECT id FROM admissions WHERE name = %s AND class_type = %s AND id != %s",
        (new_name, new_class_type, record_id),
    ).fetchone()
    if existing:
        return jsonify({"success": False, "message": "该姓名在此班型下已存在"})

    old = db.execute(
        "SELECT name, class_type FROM admissions WHERE id = %s", (record_id,),
    ).fetchone()
    if not old:
        return jsonify({"success": False, "message": "记录不存在"})

    db.execute(
        "UPDATE admissions SET name = %s, class_type = %s WHERE id = %s",
        (new_name, new_class_type, record_id),
    )
    db.commit()

    log_audit("编辑录取名单", old["name"],
              f"姓名: {old['name']}→{new_name}, 班型: {CLASS_TYPES.get(old['class_type'], {}).get('name', old['class_type'])}→{CLASS_TYPES.get(new_class_type, {}).get('name', new_class_type)}")

    return jsonify({"success": True})


@app.route("/api/admin/list/export")
@api_login_required
def admin_list_export():
    db = get_db()
    rows = db.execute(
        "SELECT name, class_type, created_at FROM admissions ORDER BY id DESC"
    ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["姓名", "班型", "添加时间"])
    for r in rows:
        writer.writerow([
            r["name"],
            CLASS_TYPES.get(r["class_type"], {}).get("name", r["class_type"] or "-"),
            r["created_at"],
        ])

    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    resp.headers["Content-Disposition"] = "attachment; filename=admissions_list.csv"
    return resp


@app.route("/api/admin/list/export_xlsx")
@api_login_required
def admin_list_export_xlsx():
    db = get_db()
    rows = db.execute(
        "SELECT name, class_type, created_at FROM admissions ORDER BY id DESC"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "录取名单"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["姓名", "班型", "添加时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        values = [
            r["name"],
            CLASS_TYPES.get(r["class_type"], {}).get("name", r["class_type"] or "-"),
            str(r["created_at"]) if r["created_at"] else "-",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = "attachment; filename=admissions_list.xlsx"
    return resp


@app.route("/api/admin/delete/<int:record_id>", methods=["DELETE"])
@api_login_required
def admin_delete(record_id):
    db = get_db()
    row = db.execute("SELECT name, class_type FROM admissions WHERE id = %s", (record_id,)).fetchone()
    db.execute("DELETE FROM admissions WHERE id = %s", (record_id,))
    db.commit()
    if row:
        log_audit("删除录取名单", row["name"], f"班型: {CLASS_TYPES.get(row['class_type'], {}).get('name', row['class_type'])}")
    return jsonify({"success": True})


@app.route("/api/admin/batch-delete", methods=["POST"])
@api_login_required
def admin_batch_delete():
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"success": False, "message": "未选择记录"})
    db = get_db()
    placeholders = ",".join(["%s"] * len(ids))
    db.execute(f"DELETE FROM admissions WHERE id IN ({placeholders})", ids)
    db.commit()
    log_audit("批量删除录取名单", "", f"删除 {len(ids)} 条记录")
    return jsonify({"success": True, "deleted": len(ids)})


@app.route("/api/admin/clear", methods=["DELETE"])
@api_login_required
def admin_clear():
    db = get_db()
    db.execute("DELETE FROM admissions")
    db.commit()
    log_audit("清空录取名单", "", "已清空全部录取名单")
    return jsonify({"success": True})


# ═══════════════════════════════════════════
# 查询日志
# ═══════════════════════════════════════════

@app.route("/api/schedule/confirm", methods=["POST"])
def schedule_confirm():
    data = request.get_json()
    name = data.get("name", "").strip()
    needs = data.get("needs", [])

    if not name:
        return jsonify({"success": False, "message": "缺少姓名"})

    needs_str = ",".join(needs) if isinstance(needs, list) else ""

    db = get_db()
    db.execute(
        """UPDATE query_logs SET needs = %s
           WHERE name = %s AND admitted = 1
           AND id = (SELECT MAX(id) FROM (SELECT id FROM query_logs WHERE name = %s AND admitted = 1) t)""",
        (needs_str, name, name),
    )
    db.commit()
    return jsonify({"success": True})


@app.route("/api/admin/logs")
@api_login_required
def admin_logs():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "").strip()
    filter_class = request.args.get("class_type", "").strip()

    db = get_db()

    conditions = []
    params = []
    if search:
        conditions.append("name LIKE %s")
        params.append(f"%{search}%")
    if filter_class:
        conditions.append("class_type = %s")
        params.append(filter_class)

    if conditions:
        where = " AND ".join(conditions)
        count = db.execute(
            f"SELECT COUNT(*) as cnt FROM query_logs WHERE {where}", params
        ).fetchone()["cnt"]
        rows = db.execute(
            f"SELECT id, name, admitted, class_type, needs, created_at FROM query_logs WHERE {where} ORDER BY id DESC LIMIT %s OFFSET %s",
            params + [per_page, (page - 1) * per_page],
        ).fetchall()
    else:
        count = db.execute("SELECT COUNT(*) as cnt FROM query_logs").fetchone()["cnt"]
        rows = db.execute(
            "SELECT id, name, admitted, class_type, needs, created_at FROM query_logs ORDER BY id DESC LIMIT %s OFFSET %s",
            (per_page, (page - 1) * per_page),
        ).fetchall()

    return jsonify({
        "total": count,
        "page": page,
        "per_page": per_page,
        "items": [dict(r) for r in rows] if rows else []
    })


@app.route("/api/admin/logs/export")
@api_login_required
def admin_logs_export():
    db = get_db()
    rows = db.execute(
        "SELECT name, admitted, class_type, needs, created_at FROM query_logs ORDER BY id DESC"
    ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["姓名", "班型", "录取状态", "培养需求", "查询时间"])
    for r in rows:
        writer.writerow([
            r["name"],
            CLASS_TYPES.get(r["class_type"], {}).get("name", r["class_type"] or "-"),
            "已录取" if r["admitted"] else "未录取",
            r["needs"] or "-",
            r["created_at"],
        ])

    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    resp.headers["Content-Disposition"] = "attachment; filename=query_logs.csv"
    return resp


@app.route("/api/admin/logs/export_xlsx")
@api_login_required
def admin_logs_export_xlsx():
    db = get_db()
    rows = db.execute(
        "SELECT name, admitted, class_type, needs, ip_address, created_at FROM query_logs ORDER BY id DESC"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "查询日志"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["姓名", "班型", "录取状态", "培养需求", "查询IP", "查询时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        ip_val = r["ip_address"] if "ip_address" in r.keys() else ""
        values = [
            r["name"],
            CLASS_TYPES.get(r["class_type"], {}).get("name", r["class_type"] or "-"),
            "已录取" if r["admitted"] else "未录取",
            r["needs"] or "-",
            ip_val or "-",
            str(r["created_at"]) if r["created_at"] else "-",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = "attachment; filename=query_logs.xlsx"
    return resp


@app.route("/api/admin/logs/batch-delete", methods=["POST"])
@api_login_required
def admin_logs_batch_delete():
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"success": False, "message": "未选择记录"})
    db = get_db()
    placeholders = ",".join(["%s"] * len(ids))
    db.execute(f"DELETE FROM query_logs WHERE id IN ({placeholders})", ids)
    db.commit()
    log_audit("批量删除查询日志", "", f"删除 {len(ids)} 条日志")
    return jsonify({"success": True, "deleted": len(ids)})


@app.route("/api/admin/logs/clear", methods=["DELETE"])
@api_login_required
def admin_logs_clear():
    db = get_db()
    db.execute("DELETE FROM query_logs")
    db.commit()
    log_audit("清空查询日志", "", "已清空全部查询日志")
    return jsonify({"success": True})


# ═══════════════════════════════════════════
# 操作日志
# ═══════════════════════════════════════════

@app.route("/api/admin/audit-logs")
@api_login_required
def admin_audit_logs():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)

    db = get_db()
    count = db.execute("SELECT COUNT(*) as cnt FROM admin_audit_log").fetchone()["cnt"]
    rows = db.execute(
        "SELECT id, action, target, detail, admin_user, created_at FROM admin_audit_log ORDER BY id DESC LIMIT %s OFFSET %s",
        (per_page, (page - 1) * per_page),
    ).fetchall()

    return jsonify({
        "total": count,
        "page": page,
        "per_page": per_page,
        "items": [dict(r) for r in rows] if rows else []
    })


@app.route("/api/admin/audit-logs/clear", methods=["DELETE"])
@api_login_required
def admin_audit_logs_clear():
    db = get_db()
    db.execute("DELETE FROM admin_audit_log")
    db.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════
# 验证码清理
# ═══════════════════════════════════════════

@app.route("/api/captcha/cleanup")
def captcha_cleanup():
    db = get_db()
    if DB_IS_MYSQL:
        db.execute("DELETE FROM captcha_store WHERE used = 0 AND created_at < DATE_SUB(NOW(), INTERVAL 10 MINUTE)")
    else:
        db.execute("DELETE FROM captcha_store WHERE used = 0 AND created_at < datetime('now', '-10 minutes')")
    db.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════
# 错误页面
# ═══════════════════════════════════════════

@app.errorhandler(404)
def page_not_found(e):
    return render_template("error.html", code=404, message="页面不存在"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("error.html", code=500, message="服务器内部错误"), 500


# ═══════════════════════════════════════════
# 启动
# ═══════════════════════════════════════════

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
